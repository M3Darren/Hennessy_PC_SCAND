import io
import os
import shutil
from datetime import datetime, timedelta
import fitz
import queue
import openpyxl
import pandas as pd
from PIL import Image
import win32com.client as win32
from get_log import m_logger, GetLog
from file_exception import FileNotFoundException, ApplicationException
from load_yaml_config import _case_scaling_ratio, _case_scanning_mode_list, \
    _case_preservation_method

# 取消图片加载的内存限制
Image.MAX_IMAGE_PIXELS = None


class LoadOperation:
    """
    @Author: M3Darren
    @Described: 加载操作，将加载的资源和Case转化为对象
    """

    _case_queue = queue.Queue()  # case对象队列
    _merge_queue = queue.Queue()  # 合并对象队列
    __image_queue = queue.Queue()  # 图片对象队列
    __pdf_queue = queue.Queue()  # pdf对象队列
    __CASE_PATH = './case.xlsx'
    __RESOURCES_PATH = None
    __PDF_RESOURCES_FLAG = False
    __IMAGE_RESOURCES_FLAG = False
    _sheet_model = None
    _sheet_name = ['JiyinScan', 'PanelScan']

    def __init__(self, resources_path, scan_mode, backup_flag):
        self.__RESOURCES_PATH = resources_path
        self._sheet_model = scan_mode
        self._scanner_model = _case_scanning_mode_list[scan_mode]
        self.backup_flag = backup_flag
        """
        Constructor
        """
        m_logger.info('read case ...')
        self.load_case_and_convert()  # 读取用例并转化
        m_logger.info('load resources ...')
        self.load_resources()
        m_logger.info('merge data ........')
        self.constructive_load_data()  # 合并数据
        m_logger.info('data integrity check ........')

    @staticmethod
    def iterator(obj):
        """
        Traversal obj
        """
        for item in obj.queue:
            print(item)

    @classmethod
    def check_resources_exist(cls):
        """
        Check resources exist
        """
        cls.close_caseFile()
        cls.clear_overthrew_logs()
        if not os.path.exists(cls.__CASE_PATH):
            raise FileNotFoundException(cls.__CASE_PATH)

    def load_case_and_convert(self):
        """
        Load case and converter.
        """
        df = pd.read_excel(self.__CASE_PATH, sheet_name=self._sheet_name[self._sheet_model])
        for row in df.to_dict(orient='records'):

            scaling_str = row[_case_scaling_ratio]
            if isinstance(scaling_str, datetime):
                scaling_str = scaling_str.strftime("%H:%M")
                row[_case_scaling_ratio] = int(scaling_str[-1])
            elif isinstance(scaling_str, str):
                row[_case_scaling_ratio] = int(scaling_str[-1])
            else:
                row[_case_scaling_ratio] = 1
            if "双" in row[self._scanner_model]:
                row[self._scanner_model] = 1
            if row[_case_preservation_method] == "PDF":
                row[_case_preservation_method] = 1
                self.__PDF_RESOURCES_FLAG = True
            else:
                self.__IMAGE_RESOURCES_FLAG = True
            self._case_queue.put(row)

    def load_resources(self):
        """
        Load resources
        """

        def get_info(file_path):
            try:
                img = Image.open(file_path)
            except Exception as e:
                raise ApplicationException(f"{file_path} Read error")
            # 获取图片的宽度和高度
            width, height = img.size
            # 获取位深度
            depth = img.mode
            # 获取DPI
            dpi = img.info.get('dpi')
            # 保存图片信息
            image_info = {
                'w_h': (width, height),
                'actual_dpi': dpi,
                'actual_bit_depth': depth,
                'result': '',
                'pil_image_obj': img,
            }
            return image_info

        path = self.__RESOURCES_PATH
        image_dir = os.listdir(path)
        if not image_dir:
            raise FileNotFoundException(image_dir, message='No image files found in the directory')
        # 通过创建时间排序
        sorted_files = sorted(
            image_dir,
            key=lambda x: os.path.getctime(os.path.join(path, x))
        )

        for filename in sorted_files:
            # 获取图片的完整路径
            file_path = os.path.join(path, filename)
            # 检查文件是否为图片格式
            if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.tiff')):
                # 打开图片
                image_info = get_info(file_path)
                self.__image_queue.put({**image_info, "filename": filename, "model": 0})
            elif filename.lower().endswith(('.pdf')):
                # 打开PDF文件
                try:
                    doc = fitz.open(file_path)
                except Exception as e:
                    raise ApplicationException(f"{file_path} Read error")
                # 遍历PDF中的每一页
                for page_num in range(len(doc)):
                    page = doc.load_page(page_num)
                    images = page.get_images(full=True)  # 获取页面中的所有图像
                    for img_index, img in enumerate(images):
                        xref = img[0]  # 图像的引用
                        base_image = doc.extract_image(xref)  # 提取图像为字节数据
                        image_bytes = base_image["image"]
                        # 使用Pillow打开图像
                        image_info = get_info(io.BytesIO(image_bytes))
                        self.__pdf_queue.put(
                            {**image_info, 'filename': filename + '@page_' + str(page_num + 1), "model": 1})

    def constructive_load_data(self):
        """
        Merge load data
        """

        def scanning_mode_judgment(queue_obj: queue.Queue):
            # 判断扫描方式为双面
            if case_item[self._scanner_model] == 1:
                if queue_obj.qsize() < 2:
                    raise ApplicationException(
                        f"The number of images is not enough")
                self._merge_queue.put((queue_obj.get(), queue_obj.get()))
            else:
                self._merge_queue.put(queue_obj.get())

        for case_item in self._case_queue.queue:
            if self.__image_queue.qsize() == 0 and self.__pdf_queue.qsize() == 0:
                raise ApplicationException(
                    f"Case exceed resources, Please check '{self.__CASE_PATH}'")
            # 判断保存方式为PDF
            if case_item[_case_preservation_method] == 1:
                scanning_mode_judgment(self.__pdf_queue)
            else:
                scanning_mode_judgment(self.__image_queue)
        if self.__image_queue.qsize() > 0 or self.__pdf_queue.qsize() > 0:
            raise ApplicationException(
                f"Resources exceed case, Please check '{self.__RESOURCES_PATH}' or {self.__CASE_PATH}")

    def write_to_caseFile_result(self, result_list, remarks_list):
        sheet_name = self._sheet_name[self._sheet_model]
        start_row = 2  # 假设从第2行开始追加（第1行通常是标题行）
        start_result_col = 1
        start_remarks_col = 2
        # 打开已存在的Excel文件
        workbook = openpyxl.load_workbook(self.__CASE_PATH)
        # 选择工作表
        sheet = workbook[sheet_name]
        # 找到要追加数据的起始单元格（例如，A2）
        start_cell_result = sheet.cell(row=start_row, column=start_result_col)
        start_cell_remarks = sheet.cell(row=start_row, column=start_remarks_col)

        # 追加数据到工作表
        for idx, value in enumerate(result_list, start=1):
            start_cell_result.offset(row=idx - 1).value = value
            if value == 'Fail':
                start_cell_remarks.offset(row=idx - 1).value = remarks_list.get()
            # 保存工作簿
        workbook.save(self.__CASE_PATH)

    @classmethod
    def close_caseFile(cls):
        """
        Close case file
        """
        excel = win32.Dispatch("Excel.Application")
        # 检查Excel是否正在运行
        if excel.Application.Workbooks.Count > 0:
            for wb in excel.Application.Workbooks:
                if wb.FullName == cls.__CASE_PATH:
                    # 尝试关闭工作簿，可能会提示保存更改
                    wb.Close(SaveChanges=True)  # SaveChanges=False 表示不保存更改
                    # break
                    raise ApplicationException("close the case file error.")
        excel.Application.Quit()

    @staticmethod
    def clear_overthrew_logs(path='log'):
        # 设置路径
        folder_path = path
        if not os.listdir(folder_path):
            return
            # 获取当前日期，并计算3天前的日期
        current_date = datetime.now()
        threshold_date = current_date - timedelta(days=3)

        # 遍历文件夹
        for folder_name in os.listdir(folder_path):
            folder_full_path = os.path.join(folder_path, folder_name)
            # 检查是否是文件夹
            if os.path.isdir(folder_full_path):
                try:
                    # 尝试将文件夹名解析为日期
                    folder_date = datetime.strptime(folder_name, "%Y_%m_%d_%H-%M-%S")
                    # 比较日期，如果早于3天前，则删除
                    if folder_date < threshold_date:
                        print(f"正在清除超出3天的log文件夹: {folder_name}")
                        shutil.rmtree(folder_full_path)  # 删除文件夹
                except ValueError:
                    # 如果文件夹名不是日期格式，则跳过
                    print(f"跳过非日期格式文件夹: {folder_name}")

    def backup_file(self, dest_path=GetLog.log_path):
        case_path = self.__CASE_PATH
        files_path = self.__RESOURCES_PATH
        shutil.copy2(case_path, dest_path)
        # 加载Excel文件
        wb = openpyxl.load_workbook(case_path)
        ws = wb.active
        # 遍历工作表中的所有行，从最后一行到第2行
        for row in range(ws.max_row, 1, -1):
            ws.delete_rows(row)
        wb.save(case_path)
        try:
            if self.backup_flag.upper() == 'N':
                print('The backup status is False. If you need backup, please use "-b Y" to specify')
            else:
                # 使用移动操作代替复制和清空
                shutil.copytree(files_path, dest_path + '/files', dirs_exist_ok=True)
                print(f"Backup {files_path} to {dest_path}")
            # os.mkdir(files_path)
        except Exception as e:
            raise ApplicationException(f"Failed to backup {files_path} to {dest_path}. Reason: {e}")
        m_logger.info(f"Cleared {files_path}")
